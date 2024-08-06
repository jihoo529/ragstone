import pandas as pd
from fastapi import HTTPException, UploadFile
import io
from typing import Dict, Any
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import logging

class RFPService:
    @staticmethod
    def read_file(file: UploadFile) -> pd.DataFrame:
        file_extension = file.filename.split('.')[-1].lower()
        content = file.file.read()

        if file_extension == 'xlsx':
            return pd.read_excel(io.BytesIO(content))
        elif file_extension == 'csv':
            return pd.read_csv(io.BytesIO(content))
        else:
            raise ValueError(f"Unsupported file format: {file_extension}. Please upload .xlsx or .csv files.")

    @staticmethod
    def prepare_queries(df: pd.DataFrame) -> Dict[str, Dict[str, str]]:
        queries = {"Questions": {}, "Answers": {}, "Sources": {}}
        for index, row in df.iterrows():
            queries["Questions"][str(index)] = row.get('Questions', '')
            queries["Answers"][str(index)] = ""  # Initialize with empty string
            queries["Sources"][str(index)] = []  # Initialize with empty list
        return queries

    @staticmethod
    def process_rfp(file: UploadFile) -> Dict[str, Any]:
        try:
            df = RFPService.read_file(file)
            queries = RFPService.prepare_queries(df)
            print("@@@@", df)
            print("####QUERIES",len(queries))

            return {
                "filename": file.filename,
                "rows": len(df),
                "columns": df.columns.tolist(),
                "queries": queries
            }
        except ValueError as ve:
            raise HTTPException(status_code=400, detail=str(ve))
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error processing file {file.filename}: {str(e)}")
    
    @staticmethod
    def json_to_xl(json_data):
        logger = logging.getLogger(__name__)
        try:
            # Create a workbook and a sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "QA Data"

            # Set up the header row
            headers = ["Questions", "Answers", "Sources"]
            ws.append(headers)

            # Style the header row
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Parse the JSON data
            questions = json_data.get("Questions", {})
            answers = json_data.get("Answers", {})
            sources = json_data.get("Sources", {})

            logger.debug(f"Questions keys: {questions.keys()}")
            logger.debug(f"Answers keys: {answers.keys()}")
            logger.debug(f"Sources keys: {sources.keys()}")

            for q_id in questions.keys():
                question = questions[q_id]
                answer = answers.get(q_id, "No answer available")
                source_list = sources.get(q_id, [])
                
                # Join the sources into a single string
                sources_str = "; ".join(source_list) if source_list else "No sources"

                ws.append([question, answer, sources_str])

            # Auto-adjust column width
            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column)  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min((max_length + 2), 100)  # Cap width at 100 for readability
                ws.column_dimensions[column].width = adjusted_width

            # Save the workbook to a bytes buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return output
        except Exception as e:
            logger.error(f"Error in json_to_xl: {str(e)}")
            raise
